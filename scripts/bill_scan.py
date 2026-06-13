import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook(
    r'C:\Users\HP\Downloads\April 2026.xlsx',
    data_only=False
)

for sheet in ['Sheet2','Bill  5th','Bill  6 To 8']:

    ws = wb[sheet]

    print("="*100)
    print(sheet)
    print("="*100)

    for row in range(1,40):

        vals=[]

        for col in range(1,10):

            v=ws.cell(row,col).value

            if v is not None:
                vals.append(f"{row},{col}={v}")

        if vals:
            print(vals)
