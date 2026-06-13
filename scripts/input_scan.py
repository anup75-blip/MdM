import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook(
    r'C:\Users\HP\Downloads\April 2026.xlsx',
    data_only=False
)

ws = wb['Sheet1']

for row in range(5,36):

    values = []

    for col in range(1,15):

        values.append(
            str(ws.cell(row,col).value)
        )

    print(row, values)
