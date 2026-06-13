import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r"C:\Users\HP\Downloads\April 2026.xlsx", data_only=False)
ws = wb["Sheet1"]

print("Scanning Sheet1 cols 29-73 (AC to BU), rows 3-5:\n")
for col in range(29, 74):
    val3 = ws.cell(3, col).value
    val5 = ws.cell(5, col).value
    if val3 is not None or val5 is not None:
        letter = get_column_letter(col)
        print(f"Col {letter:4} ({col:3}) | Row3: {val3} | Row5: {val5}")
