import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook(r'C:\Users\HP\Downloads\April 2026.xlsx', data_only=False)

for sheet in wb.sheetnames:

    ws = wb[sheet]

    print('='*100)
    print('SHEET:', sheet)
    print('='*100)

    print('ROWS:', ws.max_row)
    print('COLS:', ws.max_column)

    print('\nNON EMPTY CELLS\n')

    count = 0

    for row in ws.iter_rows():

        for cell in row:

            if cell.value is not None:

                print(f'{cell.coordinate:<8} | {cell.value}')

                count += 1

                if count >= 300:
                    break

        if count >= 300:
            break

    print('\nFORMULAS\n')

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f'{cell.coordinate} -> {cell.value}')
