"""
daily_entry.py
Reads attendance from a CSV file and fills Sheet1 in the MDM Excel workbook.

CSV format:
    Date,Class5,Class6,Class7,Class8
    01/04/2026,40,12,14,10
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# Sheet1 column indices (1-based)
COL_DATE   = 2   # B
COL_CLASS5 = 3   # C
COL_CLASS6 = 5   # E
COL_CLASS7 = 6   # F
COL_CLASS8 = 7   # G

SHEET1    = "Sheet1"
ROW_START = 5
ROW_END   = 35   # covers up to 31-day months; row 36 = monthly totals (formula row)


def load_attendance(csv_path: Path) -> pd.DataFrame:
    """
    Load and validate attendance from CSV.
    Returns a DataFrame sorted by Date with integer attendance counts.
    """
    logger.info("Loading attendance: %s", csv_path)

    if not csv_path.exists():
        raise FileNotFoundError(f"Attendance CSV not found: {csv_path}")

    df = pd.read_csv(csv_path, parse_dates=["Date"], dayfirst=True)

    required = {"Date", "Class5", "Class6", "Class7", "Class8"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"CSV missing required columns: {missing}")

    df = df.sort_values("Date").reset_index(drop=True)

    count_cols = ["Class5", "Class6", "Class7", "Class8"]
    df[count_cols] = df[count_cols].fillna(0).astype(int)

    logger.info("Loaded %d attendance records", len(df))
    return df


def fill_attendance(ws: Worksheet, df: pd.DataFrame) -> None:
    """
    Write Class5/6/7/8 attendance values into Sheet1.
    Matches rows by date from column B.
    All formula columns are untouched — Excel recalculates on open.
    """
    logger.info("Filling attendance into Sheet1")

    # Build date-string → row mapping from column B
    date_to_row: dict[str, int] = {}
    for row in range(ROW_START, ROW_END + 1):
        raw = ws.cell(row, COL_DATE).value
        if raw is None:
            continue
        key = raw.strftime("%d/%m/%Y") if hasattr(raw, "strftime") else str(raw).strip()
        date_to_row[key] = row

    filled = skipped = 0

    for _, rec in df.iterrows():
        date_key = rec["Date"].strftime("%d/%m/%Y")

        if date_key not in date_to_row:
            logger.warning("Date %s not found in Sheet1 — skipped", date_key)
            skipped += 1
            continue

        row = date_to_row[date_key]
        ws.cell(row, COL_CLASS5).value = int(rec["Class5"])
        ws.cell(row, COL_CLASS6).value = int(rec["Class6"])
        ws.cell(row, COL_CLASS7).value = int(rec["Class7"])
        ws.cell(row, COL_CLASS8).value = int(rec["Class8"])

        logger.debug(
            "Row %2d | %s | C5=%-3d C6=%-3d C7=%-3d C8=%-3d",
            row, date_key,
            rec["Class5"], rec["Class6"], rec["Class7"], rec["Class8"],
        )
        filled += 1

    logger.info("Done — filled=%d  skipped=%d", filled, skipped)


def save_workbook(wb: Workbook, output_path: Path) -> None:
    """Save workbook, creating parent directories as needed."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Workbook saved → %s", output_path)
