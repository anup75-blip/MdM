import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook

wb = load_workbook(
    r'C:\Users\HP\Downloads\April 2026.xlsx',
    data_only=False
)

for sheet in wb.sheetnames:

    ws = wb[sheet]

    print("="*100)
    print("SHEET:", sheet)
    print("="*100)

    print("\nMERGED CELLS\n")

    for rng in ws.merged_cells.ranges:
        print(rng)

    print("\nFORMULAS\n")

    count = 0

    for row in ws.iter_rows():

        for cell in row:

            if isinstance(cell.value,str) and cell.value.startswith("="):

                print(
                    f"{cell.coordinate} -> {cell.value}"
                )

                count += 1

                if count >= 200:
                    break

        if count >= 200:
            break
