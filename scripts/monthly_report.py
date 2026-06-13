"""
monthly_report.py
Copies the MDM Excel template, populates attendance, and saves the final workbook.
"""

from __future__ import annotations

import logging
import shutil
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.workbook import Workbook

from daily_entry import fill_attendance
from holiday_manager import HolidayManager

logger = logging.getLogger(__name__)

ATTENDANCE_COLS = ["Class5", "Class6", "Class7", "Class8"]


def copy_template(template_path: Path, output_path: Path) -> Workbook:
    """
    Copy Excel template to output location and return an open Workbook.
    The original template is never modified.
    """
    if not template_path.exists():
        raise FileNotFoundError(f"Template not found: {template_path}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(str(template_path), str(output_path))
    logger.info("Template copied → %s", output_path)

    wb = load_workbook(str(output_path))
    logger.info("Workbook opened. Sheets: %s", wb.sheetnames)
    return wb


def generate_report(
    wb: Workbook,
    df: pd.DataFrame,
    hm: HolidayManager,
) -> None:
    """
    Zero out attendance for non-working days, then populate Sheet1.
    All stock and expense calculations remain formula-driven.
    """
    ws = wb["Sheet1"]
    df = df.copy()

    def _to_date(val) -> date:
        return val.date() if hasattr(val, "date") else val

    # Vectorised: build mask of non-working days
    non_working = df["Date"].apply(lambda d: not hm.is_working_day(_to_date(d)))
    zeroed = int(non_working.sum())

    if zeroed:
        df.loc[non_working, ATTENDANCE_COLS] = 0
        skipped_dates = df.loc[non_working, "Date"].dt.strftime("%d/%m/%Y").tolist()
        logger.info("Zeroed %d non-working days: %s", zeroed, ", ".join(skipped_dates))

    fill_attendance(ws, df)
    logger.info("Sheet1 populated successfully")


def save_report(wb: Workbook, output_path: Path) -> None:
    """Save the populated workbook to disk."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    logger.info("Report saved → %s", output_path)


def build_output_path(base_dir: Path, year: int, month: int) -> Path:
    """Return the standard output path: output/excel/MMM_YYYY_Final.xlsx"""
    label = date(year, month, 1).strftime("%b").upper()
    return base_dir / "output" / "excel" / f"{label}_{year}_Final.xlsx"
