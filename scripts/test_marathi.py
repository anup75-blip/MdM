import sys
sys.stdout.reconfigure(encoding="utf-8")
from openpyxl import load_workbook
from marathi import to_marathi

wb = load_workbook(r"C:\Users\HP\Downloads\April 2026.xlsx", data_only=False)
ws = wb["Sheet2"]

for row in range(10, 22):
    for col in range(1, 10):
        v = ws.cell(row, col).value
        if v is not None:
            print(f"Row{row} Col{col}: {to_marathi(v)}")
