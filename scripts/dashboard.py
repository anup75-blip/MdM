"""
dashboard.py
Streamlit web dashboard for Maharashtra MDM daily attendance entry.

Run:
    streamlit run scripts/dashboard.py
"""

from __future__ import annotations

import calendar
import shutil
import sys
from datetime import date
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook

# ── Path setup ─────────────────────────────────────────────────────────────────
BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(Path(__file__).resolve().parent))

from bill_viewer import read_bill_5, read_bill_68
from daily_entry import COL_CLASS5, COL_CLASS6, COL_CLASS7, COL_CLASS8, COL_DATE, ROW_END, ROW_START
from holiday_manager import HolidayManager
from marathi import to_marathi

MASTER_FILE  = BASE_DIR / "data" / "masters" / "school_master.xlsx"
REPORTS_DIR  = BASE_DIR / "data" / "monthly_reports"
TEMPLATE     = Path(r"C:\Users\HP\Downloads\April 2026.xlsx")
HOLIDAY_FILE = BASE_DIR / "data" / "holidays" / "holiday_list.xlsx"

# ── Page config ────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="MDM Attendance",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Workbook helpers ───────────────────────────────────────────────────────────

def workbook_path(school_code: str, year: int, month: int) -> Path:
    label = date(year, month, 1).strftime("%b").upper()
    return REPORTS_DIR / str(year) / f"{month:02d}" / f"{school_code}_{label}_{year}.xlsx"


def ensure_workbook(school_code: str, year: int, month: int) -> Path:
    path = workbook_path(school_code, year, month)
    if not path.exists():
        if not TEMPLATE.exists():
            st.error(f"Template not found: {TEMPLATE}")
            st.stop()
        path.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(str(TEMPLATE), str(path))
    return path


def read_month_attendance(wb_path: Path) -> pd.DataFrame:
    wb = load_workbook(str(wb_path), data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for row in range(ROW_START, ROW_END + 1):
        raw = ws.cell(row, COL_DATE).value
        if raw is None:
            continue
        date_str = raw.strftime("%d/%m/%Y") if hasattr(raw, "strftime") else str(raw).strip()
        rows.append({
            "Date":   date_str,
            "Class5": ws.cell(row, COL_CLASS5).value or 0,
            "Class6": ws.cell(row, COL_CLASS6).value or 0,
            "Class7": ws.cell(row, COL_CLASS7).value or 0,
            "Class8": ws.cell(row, COL_CLASS8).value or 0,
        })
    wb.close()
    return pd.DataFrame(rows)


def read_sheet_preview(wb_path: Path, sheet_name: str, max_row: int = 40, max_col: int = 12) -> pd.DataFrame:
    """
    Read any sheet as a plain DataFrame.
    String cells are converted from Kruti Dev to Unicode Marathi.
    Numeric cells are shown as-is.
    """
    wb = load_workbook(str(wb_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return pd.DataFrame()
    ws = wb[sheet_name]
    data = []
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True):
        converted = []
        for cell in row:
            if cell is None:
                converted.append("")
            elif isinstance(cell, str):
                converted.append(to_marathi(cell))
            else:
                converted.append(str(cell))
        data.append(converted)
    wb.close()
    df = pd.DataFrame(data)
    df = df[df.apply(lambda r: r.str.strip().any(), axis=1)].reset_index(drop=True)
    return df


def save_day_attendance(wb_path: Path, date_str: str, c5: int, c6: int, c7: int, c8: int) -> bool:
    wb = load_workbook(str(wb_path))
    ws = wb["Sheet1"]
    saved = False
    for row in range(ROW_START, ROW_END + 1):
        raw = ws.cell(row, COL_DATE).value
        if raw is None:
            continue
        key = raw.strftime("%d/%m/%Y") if hasattr(raw, "strftime") else str(raw).strip()
        if key == date_str:
            ws.cell(row, COL_CLASS5).value = c5
            ws.cell(row, COL_CLASS6).value = c6
            ws.cell(row, COL_CLASS7).value = c7
            ws.cell(row, COL_CLASS8).value = c8
            saved = True
            break
    if saved:
        wb.save(str(wb_path))
    wb.close()
    return saved


@st.cache_data(show_spinner=False)
def load_school_master() -> pd.DataFrame:
    if not MASTER_FILE.exists():
        return pd.DataFrame(columns=["SchoolName", "SchoolCode"])
    return pd.read_excel(str(MASTER_FILE), dtype=str)


# ── App ────────────────────────────────────────────────────────────────────────

st.title("Mid-Day Meal — Attendance Dashboard")
st.caption("Maharashtra MDM Scheme | Daily Entry by School")

# ── Sidebar ────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.header("School & Period")

    schools_df = load_school_master()
    if schools_df.empty:
        st.error("No schools in school_master.xlsx")
        st.markdown(f"`{MASTER_FILE}`")
        st.stop()

    selected_name = st.selectbox("School", schools_df["SchoolName"].tolist())
    school_row    = schools_df[schools_df["SchoolName"] == selected_name].iloc[0]
    school_code   = str(school_row["SchoolCode"]).strip()

    today = date.today()
    month_map = {date(2026, m, 1).strftime("%B"): m for m in range(1, 13)}
    selected_month_name = st.selectbox("Month", list(month_map.keys()), index=today.month - 1)
    selected_month = month_map[selected_month_name]
    selected_year  = st.number_input("Year", value=2026, min_value=2020, max_value=2030, step=1)

    wb_path = ensure_workbook(school_code, selected_year, selected_month)

    st.divider()
    st.caption(f"Code: `{school_code}`")
    st.caption(f"File: `{wb_path.name}`")

    st.divider()
    st.subheader("Download Report")
    with open(wb_path, "rb") as f:
        st.download_button(
            label="Download Excel",
            data=f.read(),
            file_name=wb_path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    st.caption("Opens in Excel with all formulas calculated.")

# ── Holiday manager ────────────────────────────────────────────────────────────
hm = HolidayManager(year=selected_year)
hm.load_school_holidays(HOLIDAY_FILE)

# ── Tabs ───────────────────────────────────────────────────────────────────────
tab_entry, tab_s1, tab_s2, tab_bill5, tab_bill68 = st.tabs([
    "Enter Attendance",
    "Sheet 1 — Daily Register",
    "Sheet 2 — Monthly Summary",
    "Bill — Class 5",
    "Bill — Class 6 to 8",
])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — ATTENDANCE ENTRY
# ══════════════════════════════════════════════════════════════════════════════
with tab_entry:
    st.subheader(f"Enter Attendance — {selected_name}")

    days_in_month = calendar.monthrange(selected_year, selected_month)[1]
    min_date  = date(selected_year, selected_month, 1)
    max_date  = date(selected_year, selected_month, days_in_month)
    def_date  = today if (min_date <= today <= max_date) else min_date

    col_date, col_status = st.columns([1, 2])
    with col_date:
        selected_date = st.date_input("Date", value=def_date, min_value=min_date, max_value=max_date)

    date_str    = selected_date.strftime("%d/%m/%Y")
    holiday_rsn = hm.holiday_name(selected_date)

    with col_status:
        st.write("")
        if holiday_rsn:
            st.error(f"Non-working day: {holiday_rsn}")
        else:
            st.success("Working day")

    # Read existing values
    month_df  = read_month_attendance(wb_path)
    match     = month_df[month_df["Date"] == date_str]
    has_entry = not match.empty

    d_c5 = int(match["Class5"].values[0]) if has_entry else 0
    d_c6 = int(match["Class6"].values[0]) if has_entry else 0
    d_c7 = int(match["Class7"].values[0]) if has_entry else 0
    d_c8 = int(match["Class8"].values[0]) if has_entry else 0

    if holiday_rsn:
        d_c5 = d_c6 = d_c7 = d_c8 = 0

    with st.form("attendance_form", clear_on_submit=False):
        a, b, c, d_ = st.columns(4)
        c5 = a.number_input("Class 5", min_value=0, max_value=999, value=d_c5, step=1)
        c6 = b.number_input("Class 6", min_value=0, max_value=999, value=d_c6, step=1)
        c7 = c.number_input("Class 7", min_value=0, max_value=999, value=d_c7, step=1)
        c8 = d_.number_input("Class 8", min_value=0, max_value=999, value=d_c8, step=1)

        m1, m2, m3 = st.columns(3)
        m1.metric("Class 5 Total",   c5)
        m2.metric("Class 6-8 Total", c6 + c7 + c8)
        m3.metric("Grand Total",     c5 + c6 + c7 + c8)

        submitted = st.form_submit_button("Save Attendance", use_container_width=True, type="primary")

    if submitted:
        if holiday_rsn and (c5 + c6 + c7 + c8) > 0:
            st.warning(f"Saving non-zero attendance on a {holiday_rsn}.")
        ok = save_day_attendance(wb_path, date_str, c5, c6, c7, c8)
        if ok:
            st.success(f"Saved — {date_str} | Cl5={c5}  Cl6={c6}  Cl7={c7}  Cl8={c8}")
            st.cache_data.clear()
        else:
            st.error(f"Date {date_str} not found in workbook — check template month.")

    # Month overview
    st.divider()
    month_label = date(selected_year, selected_month, 1).strftime("%B %Y")
    st.subheader(f"{month_label} Overview")

    overview = read_month_attendance(wb_path)
    if not overview.empty:
        def day_status(ds: str) -> str:
            try:
                d2 = date(selected_year, selected_month, int(ds.split("/")[0]))
                r  = hm.holiday_name(d2)
                return r if r else "Working"
            except Exception:
                return ""

        overview["Status"] = overview["Date"].apply(day_status)
        overview["Total"]  = overview[["Class5", "Class6", "Class7", "Class8"]].sum(axis=1)
        overview["Filled"] = overview["Total"].apply(lambda x: "Yes" if x > 0 else "—")

        st.dataframe(
            overview[["Date", "Status", "Class5", "Class6", "Class7", "Class8", "Total", "Filled"]],
            use_container_width=True, hide_index=True,
        )

        working = overview[overview["Status"] == "Working"]
        filled  = working[working["Total"] > 0]
        s1, s2, s3, s4 = st.columns(4)
        s1.metric("Working Days",    len(working))
        s2.metric("Days Filled",     len(filled))
        s3.metric("Class 5 Total",   int(overview["Class5"].sum()))
        s4.metric("Class 6-8 Total", int(overview[["Class6", "Class7", "Class8"]].sum().sum()))

        st.divider()
        dcol, _ = st.columns([1, 2])
        with dcol:
            with open(wb_path, "rb") as f:
                st.download_button(
                    label=f"Download {wb_path.stem}.xlsx",
                    data=f.read(),
                    file_name=wb_path.name,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — SHEET 1
# ══════════════════════════════════════════════════════════════════════════════
with tab_s1:
    st.subheader("Sheet 1 — Daily Attendance Register")
    st.caption("Columns A–N shown (date, attendance, menu). Stock columns are formula-driven.")

    overview2 = read_month_attendance(wb_path)
    if not overview2.empty:
        overview2["Class 6-8"] = overview2[["Class6", "Class7", "Class8"]].sum(axis=1)
        overview2["Total"]     = overview2["Class5"] + overview2["Class 6-8"]

        def day_status2(ds: str) -> str:
            try:
                d2 = date(selected_year, selected_month, int(ds.split("/")[0]))
                r  = hm.holiday_name(d2)
                return r if r else "Working"
            except Exception:
                return ""

        overview2["Day Status"] = overview2["Date"].apply(day_status2)

        st.dataframe(
            overview2[["Date", "Day Status", "Class5", "Class6", "Class7", "Class8", "Class 6-8", "Total"]].rename(columns={
                "Class5": "Cl 5 (C)",
                "Class6": "Cl 6 (E)",
                "Class7": "Cl 7 (F)",
                "Class8": "Cl 8 (G)",
            }),
            use_container_width=True,
            hide_index=True,
        )

        r1, r2, r3 = st.columns(3)
        r1.metric("Total Class 5",   int(overview2["Class5"].sum()))
        r2.metric("Total Class 6-8", int(overview2["Class 6-8"].sum()))
        r3.metric("Grand Total",     int(overview2["Total"].sum()))
    else:
        st.info("No data yet. Enter attendance in the first tab.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — SHEET 2
# ══════════════════════════════════════════════════════════════════════════════
with tab_s2:
    st.subheader("Sheet 2 — मासिक सारांश (Monthly Summary)")

    from bill_viewer import FOOD_ITEMS, _val
    from openpyxl import load_workbook as _lw

    _wb2 = _lw(str(wb_path), data_only=True)
    if "Sheet2" in _wb2.sheetnames:
        _ws2 = _wb2["Sheet2"]

        # Header metrics
        _mc1, _mc2, _mc3 = st.columns(3)
        _mc1.metric("विद्यार्थी इ.5 (Students Cl5)", int(_val(_ws2, 5, 7) or 0))
        _mc2.metric("विद्यार्थी इ.6-8 (Students Cl6-8)", int(_val(_ws2, 6, 7) or 0))
        _mc3.metric("कामाचे दिवस (Working Days)", int(_val(_ws2, 7, 7) or 0))

        st.divider()
        st.markdown("**वस्तू साठा — इ.6 ते 8 (Food Stock Class 6-8)**")

        _rows2 = []
        for serial, row, name, *_ in FOOD_ITEMS:
            _rows2.append({
                "क्र":               serial,
                "वस्तूचे नाव":       name,
                "मागिल शिल्लक":     _val(_ws2, row, 3),
                "प्राप्त वस्तू":      _val(_ws2, row, 4),
                "एकूण (3+4)":        _val(_ws2, row, 5),
                "वापरलेल्या वस्तू":  _val(_ws2, row, 6),
                "शिल्लक (5+6)":      _val(_ws2, row, 7),
                "मागणी":             _val(_ws2, row, 8),
            })
        st.dataframe(pd.DataFrame(_rows2), use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("**खर्चाचे विवरण (Expense Summary)**")
        _exp_labels = {23: "भाजिपाला", 24: "पूरक आहार", 25: "इंधन"}
        _exp_rows = []
        for _r, _lbl in _exp_labels.items():
            _exp_rows.append({
                "तपशील":           _lbl,
                "इ.6-8 दर (₹)":   _val(_ws2, _r, 5),
                "इ.6-8 रक्कम (₹)": _val(_ws2, _r, 6),
                "इ.5 दर (₹)":     _val(_ws2, _r, 9),
                "इ.5 रक्कम (₹)":  _val(_ws2, _r, 10),
            })
        st.dataframe(pd.DataFrame(_exp_rows), use_container_width=True, hide_index=True)
    else:
        st.info("Sheet2 not found in workbook.")
    _wb2.close()

    st.caption("संख्या — Excel मध्ये उघडल्यावर सूत्रे अद्ययावत होतात.")


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 — BILL CLASS 5
# ══════════════════════════════════════════════════════════════════════════════
with tab_bill5:
    st.subheader("Bill — Class 5 | इ.5 चे विधेयक")

    _b5 = read_bill_5(wb_path)
    if not _b5:
        st.warning("Bill 5th sheet not found in workbook.")
    else:
        _a5, _b5c, _c5 = st.columns(3)
        _a5.metric("विद्यार्थी संख्या", int(_b5["students"]))
        _b5c.metric("एकूण उपस्थिती",   int(_b5["attendance"]))
        _c5.metric("कामाचे दिवस",       int(_b5["working_days"]))

        st.divider()
        st.markdown("**वस्तू साठा विवरण — इ.5**")

        _stock5 = _b5["stock"].copy()
        for _col in ["मागिल शिल्लक", "प्राप्त वस्तू", "एकूण (3+4)", "वापरलेल्या वस्तू", "शिल्लक (5+6)", "पुढील मागणी"]:
            _stock5[_col] = _stock5[_col].apply(lambda x: f"{x:.3f}" if x else "—")
        st.dataframe(_stock5, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("**खर्चाचे विवरण — इ.5**")

        _exp5 = _b5["expense"].copy()
        _exp5["मुलांची संख्या"] = _exp5["मुलांची संख्या"].apply(
            lambda x: int(x) if isinstance(x, (int, float)) and x != "" else x
        )
        _exp5["दर (₹)"]    = _exp5["दर (₹)"].apply(
            lambda x: f"₹{x:.2f}" if isinstance(x, (int, float)) and x != "" else x
        )
        _exp5["रक्कम (₹)"] = _exp5["रक्कम (₹)"].apply(
            lambda x: f"₹{x:.2f}" if isinstance(x, (int, float)) and x != "" else x
        )
        st.dataframe(_exp5, use_container_width=True, hide_index=True)

        st.caption(
            f"एकूण दर: ₹{_b5['total_rate']:.2f} | एकूण रक्कम: ₹{_b5['total_amount']:.2f}  "
            "— Excel मध्ये उघडल्यावर सूत्रे अद्ययावत होतात."
        )


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 — BILL CLASS 6-8
# ══════════════════════════════════════════════════════════════════════════════
with tab_bill68:
    st.subheader("Bill — Class 6 to 8 | इ.6 ते 8 चे विधेयक")

    _b68 = read_bill_68(wb_path)
    if not _b68:
        st.warning("Bill 6 To 8 sheet not found in workbook.")
    else:
        _a68, _b68c, _c68 = st.columns(3)
        _a68.metric("विद्यार्थी संख्या", int(_b68["students"]))
        _b68c.metric("एकूण उपस्थिती",   int(_b68["attendance"]))
        _c68.metric("कामाचे दिवस",       int(_b68["working_days"]))

        st.divider()
        st.markdown("**वस्तू साठा विवरण — इ.6 ते 8**")

        _stock68 = _b68["stock"].copy()
        for _col in ["मागिल शिल्लक", "प्राप्त वस्तू", "एकूण (3+4)", "वापरलेल्या वस्तू", "शिल्लक (5+6)", "पुढील मागणी"]:
            _stock68[_col] = _stock68[_col].apply(lambda x: f"{x:.3f}" if x else "—")
        st.dataframe(_stock68, use_container_width=True, hide_index=True)

        st.divider()
        st.markdown("**खर्चाचे विवरण — इ.6 ते 8**")

        _exp68 = _b68["expense"].copy()
        _exp68["मुलांची संख्या"] = _exp68["मुलांची संख्या"].apply(
            lambda x: int(x) if isinstance(x, (int, float)) and x != "" else x
        )
        _exp68["दर (₹)"]    = _exp68["दर (₹)"].apply(
            lambda x: f"₹{x:.2f}" if isinstance(x, (int, float)) and x != "" else x
        )
        _exp68["रक्कम (₹)"] = _exp68["रक्कम (₹)"].apply(
            lambda x: f"₹{x:.2f}" if isinstance(x, (int, float)) and x != "" else x
        )
        st.dataframe(_exp68, use_container_width=True, hide_index=True)

        st.caption(
            f"एकूण दर: ₹{_b68['total_rate']:.2f} | एकूण रक्कम: ₹{_b68['total_amount']:.2f}  "
            "— Excel मध्ये उघडल्यावर सूत्रे अद्ययावत होतात."
        )
