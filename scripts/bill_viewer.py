"""
bill_viewer.py
Reads bill sheets and returns structured data with proper Unicode Marathi labels.
"""

from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# ── Food items as they appear in Sheet2 (rows 12-21, 10 items) ────────────────
# Used for Sheet2 display tab
FOOD_ITEMS = [
    ( 1, 12, "तांदूळ"),
    ( 2, 13, "मटकी"),          # has #REF! in cols C/D — values may be 0
    ( 3, 14, "मसूर दाळ"),
    ( 4, 15, "हरभरा"),
    ( 5, 16, "वाटाणा"),
    ( 6, 17, "मोहरी"),
    ( 7, 18, "हळद"),
    ( 8, 19, "कांदा लसूण मसाला"),
    ( 9, 20, "मीठ"),
    (10, 21, "सोया तेल"),
]

# ── Food items as they appear in Bill sheets (rows 12-20, 9 items) ─────────────
# Bill sheets skip serial 2 (Matki has #REF! so it is absent)
BILL_FOOD_ITEMS = [
    ( 1, 12, "तांदूळ"),
    ( 3, 13, "मसूर दाळ"),
    ( 4, 14, "मटकी / चवळी"),
    ( 5, 15, "हरभरा"),
    ( 6, 16, "मोहरी"),
    ( 7, 17, "हळद"),
    ( 8, 18, "कांदा लसूण मसाला"),
    ( 9, 19, "मीठ"),
    (10, 20, "सोया तेल"),
]

# ── Column indices in Bill sheets (1-based) ────────────────────────────────────
COL_OPENING  = 3   # मागिल शिल्लक
COL_RECEIVED = 4   # प्राप्त वस्तू
COL_TOTAL    = 5   # एकूण (3+4)
COL_USED     = 6   # वापरलेल्या
COL_BALANCE  = 7   # शिल्लक
COL_DEMAND   = 8   # पुढील मागणी

# ── Expense rows — same layout for both Bill 5th and Bill 6-8 ─────────────────
# Row 21 = expense section header ("विवरण | मुलांची संख्या | दर | रक्कम")
# Rows 22-24 = expense data rows
# Row 25 = एकूण total
EXPENSE_ROWS = [
    (22, "भाजिपाला"),
    (23, "पूरक आहार"),
    (24, "इंधन"),
]
EXPENSE_TOTAL_ROW = 25


def _val(ws, row: int, col: int, default=0):
    """Read a cell value safely, returning default if None or unparseable."""
    v = ws.cell(row, col).value
    if v is None:
        return default
    if isinstance(v, str) and (v.startswith("=") or v.startswith("#")):
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


def read_bill(wb_path: Path, sheet_name: str) -> dict:
    """
    Read a bill sheet and return structured data with proper Unicode Marathi labels.
    Uses data_only=True so openpyxl returns cached formula values.
    """
    wb = load_workbook(str(wb_path), data_only=True)
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {}

    ws = wb[sheet_name]

    # ── Header metrics ─────────────────────────────────────────────────────────
    # Try several cells where enrollment/attendance/working-days can appear
    students     = _val(ws, 5, 5) or _val(ws, 5, 4)
    attendance   = _val(ws, 6, 5) or _val(ws, 6, 4)
    working_days = _val(ws, 5, 9) or _val(ws, 6, 9) or _val(ws, 7, 9)

    # ── Food stock table ───────────────────────────────────────────────────────
    stock_rows = []
    for serial, row, name in BILL_FOOD_ITEMS:
        stock_rows.append({
            "क्र":              serial,
            "वस्तूचे नाव":      name,
            "मागिल शिल्लक":    _val(ws, row, COL_OPENING),
            "प्राप्त वस्तू":     _val(ws, row, COL_RECEIVED),
            "एकूण (3+4)":       _val(ws, row, COL_TOTAL),
            "वापरलेल्या वस्तू": _val(ws, row, COL_USED),
            "शिल्लक (5+6)":     _val(ws, row, COL_BALANCE),
            "पुढील मागणी":      _val(ws, row, COL_DEMAND),
        })

    stock_df = pd.DataFrame(stock_rows)

    # ── Expense section ────────────────────────────────────────────────────────
    expense_data = []
    for row, label in EXPENSE_ROWS:
        qty  = _val(ws, row, 4)
        rate = _val(ws, row, 5)
        amt  = _val(ws, row, 6)
        expense_data.append({
            "तपशील":           label,
            "मुलांची संख्या":  qty,
            "दर (₹)":          rate,
            "रक्कम (₹)":       amt,
        })

    total_rate = _val(ws, EXPENSE_TOTAL_ROW, 5)
    total_amt  = _val(ws, EXPENSE_TOTAL_ROW, 6)
    expense_data.append({
        "तपशील":           "एकूण",
        "मुलांची संख्या":  "",
        "दर (₹)":          total_rate,
        "रक्कम (₹)":       total_amt,
    })

    expense_df = pd.DataFrame(expense_data)
    wb.close()

    return {
        "students":     students,
        "attendance":   attendance,
        "working_days": working_days,
        "stock":        stock_df,
        "expense":      expense_df,
        "total_rate":   total_rate,
        "total_amount": total_amt,
    }


def read_bill_68(wb_path: Path) -> dict:
    return read_bill(wb_path, "Bill  6 To 8")


def read_bill_5(wb_path: Path) -> dict:
    return read_bill(wb_path, "Bill  5th")
