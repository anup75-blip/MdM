"""
excel_export.py
Generate Excel government report in memory from Supabase attendance data.
No disk writes — returns bytes that Streamlit serves as a download.
"""
from __future__ import annotations

import calendar
import io
from datetime import date
from pathlib import Path

from openpyxl import load_workbook

# Sheet1 column indices (same as daily_entry.py)
COL_DATE   = 2   # B
COL_CLASS5 = 3   # C
COL_CLASS6 = 5   # E
COL_CLASS7 = 6   # F
COL_CLASS8 = 7   # G
COL_J_DATE = 10  # J (duplicate date column in template)
ROW_START  = 5
ROW_END    = 35


def generate_excel_buffer(
    template_path: Path,
    attendance_records: list[dict],
    year: int,
    month: int,
) -> bytes:
    """
    Build an in-memory Excel report from the government template.

    attendance_records: list of dicts — each has keys:
        date   (str 'YYYY-MM-DD' or date object)
        class5, class6, class7, class8  (int)

    Returns xlsx bytes.  Caller passes to st.download_button(data=...).
    """
    # Load template into a memory buffer (never writes to disk)
    buf = io.BytesIO()
    with open(template_path, "rb") as fh:
        buf.write(fh.read())
    buf.seek(0)

    wb = load_workbook(buf)
    ws = wb["Sheet1"]

    # ── Fix dates for requested month ──────────────────────────────────────────
    days_in_month = calendar.monthrange(year, month)[1]
    for day in range(1, 32):
        row = ROW_START + day - 1
        if day <= days_in_month:
            d = date(year, month, day)
            ws.cell(row, COL_DATE).value   = d
            ws.cell(row, COL_J_DATE).value = d
        else:
            ws.cell(row, COL_DATE).value   = None
            ws.cell(row, COL_J_DATE).value = None

    # ── Build lookup: date → record ────────────────────────────────────────────
    att_map: dict[date, dict] = {}
    for rec in attendance_records:
        d = rec["date"]
        if isinstance(d, str):
            d = date.fromisoformat(d)
        att_map[d] = rec

    # ── Write attendance numbers into Sheet1 ───────────────────────────────────
    for day in range(1, days_in_month + 1):
        row = ROW_START + day - 1
        rec = att_map.get(date(year, month, day))
        if rec:
            ws.cell(row, COL_CLASS5).value = int(rec["class5"])
            ws.cell(row, COL_CLASS6).value = int(rec["class6"])
            ws.cell(row, COL_CLASS7).value = int(rec["class7"])
            ws.cell(row, COL_CLASS8).value = int(rec["class8"])
        else:
            # Leave cells blank (not zero) so teacher can see unfilled days
            ws.cell(row, COL_CLASS5).value = None
            ws.cell(row, COL_CLASS6).value = None
            ws.cell(row, COL_CLASS7).value = None
            ws.cell(row, COL_CLASS8).value = None

    # ── Save to output buffer ─────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    wb.close()
    return out.getvalue()
